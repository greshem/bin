#coding=utf-8

import openpyxl
import sys;
reload(sys)
sys.setdefaultencoding('utf-8')

work_book = openpyxl.load_workbook("boyan.xlsx")
#print work_book.get_sheet_names()
print work_book.sheetnames;
# >>> ['Sheet1', 'Sheet2', 'Sheet3']

#print work_book.get_active_sheet().title;
print work_book.active.title;
#work_sheet=work_book.active;
work_sheet=work_book.get_sheet_by_name("Sheet1")

print "max_row {t}\n".format(t=work_sheet.max_row)
print "max_colum {t}\n".format(t=work_sheet.max_column)

for r in xrange(1, work_sheet.max_row):
    array=[ work_sheet.cell(row=r, column=c).value  for c in xrange(1, work_sheet.max_column)]
    #print array;
    for each in array:
        try:
            print "%s|"%(each),
        except: 
            info=sys.exc_info()
            print info[0],":",info[1]
    print 
        #val = work_sheet.cell(row=r, column=c).value
        #if val != None:
        #    print r,c, val
